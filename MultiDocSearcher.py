import tkinter as tk
from tkinter import ttk, filedialog, Menu
from docx import Document
from openpyxl import load_workbook
import re
import os

class DocumentSearchApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Document Search Tool")

        self.file_path = ""
        self.directory_path = ""
        self.result_text = None
        self.scrollbar = None
        self.selected_text = None  # Save selected text

        self.create_widgets()

    def create_widgets(self):
        # Create buttons for selecting file and directory
        self.select_file_button = ttk.Button(self.root, text="Select Document", command=self.select_document)
        self.select_directory_button = ttk.Button(self.root, text="Select Directory", command=self.select_directory)
        self.select_file_button.pack(pady=10, padx=10, anchor=tk.W)
        self.select_directory_button.pack(pady=10, padx=10, anchor=tk.W)

        # Information labels for displaying document and directory paths
        self.file_info_label = ttk.Label(self.root, text="Document Path:")
        self.directory_info_label = ttk.Label(self.root, text="Directory Path:")
        self.file_info_label.pack(pady=0, padx=10, anchor=tk.W)
        self.directory_info_label.pack(pady=0, padx=10, anchor=tk.W)

        # Create search entry and button
        self.search_entry = ttk.Entry(self.root, width=30)
        self.search_button = ttk.Button(self.root, text="Search", command=self.search_files_and_locate)

        # Create text box for displaying search results
        self.result_text = tk.Text(self.root, wrap=tk.WORD, height=10, width=50)
        self.result_text.config(state=tk.DISABLED)  # Set initially as non-editable

        # Create scrollbar
        self.scrollbar = tk.Scrollbar(self.root, command=self.result_text.yview)

        # Bind Enter key to the search function
        self.root.bind('<Return>', lambda event=None: self.search_files_and_locate())

        # Bind left-click and right-click events to the result_text box
        self.result_text.bind("<ButtonRelease-1>", self.left_click)
        self.result_text.bind("<ButtonRelease-3>", self.right_click)

        # Place components
        self.search_entry.pack(pady=10, padx=10, anchor=tk.W)
        self.search_button.pack(pady=10, padx=10, anchor=tk.W)
        self.result_text.pack(pady=10, padx=10, anchor=tk.W, expand=True, fill=tk.BOTH)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Set column and row weights for resizing
        for i in range(2):  # Set for the first two columns
            self.root.columnconfigure(i, weight=1)
        for i in range(5):  # Set for the first five rows
            self.root.rowconfigure(i, weight=1)

    def select_document(self):
        # Open file dialog to select a document
        self.file_path = filedialog.askopenfilename(filetypes=[("Word Documents", "*.docx"), ("Excel Documents", "*.xlsx")])

        # Update document path information
        if self.file_path:
            self.file_info_label.config(text="Document Path: " + self.file_path)
            self.directory_path = ""  # Clear directory path

    def select_directory(self):
        # Open folder dialog to select a directory
        self.directory_path = filedialog.askdirectory(title="Select Target Directory")

        # Update directory path information
        if self.directory_path:
            self.directory_info_label.config(text="Directory Path: " + self.directory_path)
            self.file_path = ""  # Clear document path

    def search_files_and_locate(self):
        if not self.file_path and not self.directory_path:
            # If both file path and directory path are empty, prompt the user to choose a document or directory
            self.result_text.config(state=tk.NORMAL)
            self.result_text.insert(tk.END, "Please select a document or directory first.\n")
            self.result_text.config(state=tk.DISABLED)
            return

        search_keyword = re.escape(self.search_entry.get())

        # Clear the text box
        self.result_text.config(state=tk.NORMAL)
        self.result_text.delete(1.0, tk.END)

        if self.file_path:
            # Search in the specified document
            self.search_file_content(self.file_path, search_keyword)
        elif self.directory_path:
            # Recursively search files in the specified directory
            self.search_files_in_directory(self.directory_path, search_keyword)

        # Set the text box to read-only
        self.result_text.config(state=tk.DISABLED)

    def search_files_in_directory(self, directory, search_keyword):
        """
        Recursively search for files containing the specified keyword in the given directory.
        """
        for root, dirs, files in os.walk(directory):
            for file in files:
                file_path = os.path.join(root, file)
                if file_path.lower().endswith(('.docx', '.xlsx')):
                    self.search_file_content(file_path, search_keyword)

    def search_file_content(self, file_path, search_keyword):
        """
        Search for the keyword in the specified file and return True if found.
        """
        found = False  # Flag to indicate if matching content is found in the file

        if file_path.lower().endswith('.docx'):
            document = Document(file_path)
            for paragraph in document.paragraphs:
                if re.search(search_keyword, paragraph.text, re.IGNORECASE):
                    found = True
                    # If matching file content is found, display the result in the text box
                    if not self.result_text.get("1.0", tk.END).count(f"Match found in file '{file_path}':"):
                        self.result_text.insert(tk.END, f"Match found in file '{file_path}':\n")
                    self.result_text.insert(tk.END, f"{paragraph.text}\n\n")

        elif file_path.lower().endswith('.xlsx'):
            try:
                wb = load_workbook(file_path)
            except Exception as e:
                # Handle exceptions
                self.result_text.insert(tk.END, f"Error reading Excel document: {str(e)}\n")
                return False

            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows():
                    row_data = [str(cell.value) for cell in row if cell.value]
                    row_text = '\t'.join(row_data)
                    if re.search(search_keyword, row_text, re.IGNORECASE):
                        found = True
                        # If matching file content is found, display the result in the text box
                        if not self.result_text.get("1.0", tk.END).count(f"Match found in file '{file_path}':"):
                            self.result_text.insert(tk.END, f"Match found in file '{file_path}':\n")
                        self.result_text.insert(tk.END, f"{row_text}\n\n")

        return found

    def left_click(self, event):
        # Check if there is selected text
        if self.result_text.tag_ranges(tk.SEL):
            # Get the selected text
            start_index = self.result_text.index(tk.SEL_FIRST)
            end_index = self.result_text.index(tk.SEL_LAST)
            self.selected_text = self.result_text.get(start_index, end_index)

    def right_click(self, event):
        # Create a right-click menu
        context_menu = Menu(self.root, tearoff=0)
        context_menu.add_command(label="Open File", command=self.open_file)

        # Display the menu at the mouse position
        context_menu.post(event.x_root, event.y_root)

    def open_file(self):
        if not self.selected_text:
            return

        # Get the file path containing the selected text
        file_path = self.find_file_containing_text(self.selected_text)

        if file_path:
            # Open the file
            try:
                os.startfile(file_path)
            except Exception as e:
                self.result_text.insert(tk.END, f"Error opening file: {str(e)}\n")

    def find_file_containing_text(self, search_text):
        """
        Recursively search for files containing the specified text in the given directory,
        and return the file path if found.
        """
        for root, dirs, files in os.walk(self.directory_path):
            for file in files:
                file_path = os.path.join(root, file)
                if file_path.lower().endswith(('.docx', '.xlsx')):
                    if self.search_file_content(file_path, re.escape(search_text)):
                        return file_path
        return None

if __name__ == "__main__":
    root = tk.Tk()
    app = DocumentSearchApp(root)
    root.mainloop()
